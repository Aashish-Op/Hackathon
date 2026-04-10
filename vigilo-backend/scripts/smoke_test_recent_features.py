import os
import uuid
import io
from copy import deepcopy
from datetime import datetime, timedelta, timezone
from typing import Any


# Ensure config loads even when local .env has blanks.
os.environ.setdefault("SUPABASE_URL", "https://example.supabase.co")
os.environ.setdefault("SUPABASE_KEY", "dummy-service-role-key")
os.environ.setdefault("SUPABASE_JWT_SECRET", "dummy-jwt-secret")
os.environ.setdefault("OPENAI_API_KEY", "dummy-openai-key")

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.auth.dependencies import get_current_user, require_student, require_tpc_admin
from app.main import app
import app.routers.ai as ai_router
import app.routers.exports as exports_router
import app.routers.interventions as interventions_router
import app.routers.notifications as notifications_router


class FakeResponse:
    def __init__(self, data: list[dict[str, Any]]):
        self.data = data


class FakeQuery:
    def __init__(self, db: dict[str, list[dict[str, Any]]], table_name: str):
        self.db = db
        self.table_name = table_name
        self.op = "select"
        self.select_fields: list[str] | None = None
        self.filters: list[tuple[str, str, Any]] = []
        self.order_by: tuple[str, bool] | None = None
        self.range_bounds: tuple[int, int] | None = None
        self.limit_count: int | None = None
        self.update_payload: dict[str, Any] | None = None
        self.insert_payload: list[dict[str, Any]] | None = None

    def _rows(self) -> list[dict[str, Any]]:
        return self.db.setdefault(self.table_name, [])

    def select(self, fields: str):
        self.op = "select"
        if fields.strip() == "*":
            self.select_fields = None
        else:
            self.select_fields = [item.strip() for item in fields.split(",")]
        return self

    def eq(self, column: str, value: Any):
        self.filters.append(("eq", column, value))
        return self

    def in_(self, column: str, values: list[Any]):
        self.filters.append(("in", column, list(values)))
        return self

    def gte(self, column: str, value: Any):
        self.filters.append(("gte", column, value))
        return self

    def lte(self, column: str, value: Any):
        self.filters.append(("lte", column, value))
        return self

    def is_(self, column: str, value: Any):
        self.filters.append(("is", column, value))
        return self

    def order(self, column: str, desc: bool = False):
        self.order_by = (column, desc)
        return self

    def range(self, start: int, end: int):
        self.range_bounds = (start, end)
        return self

    def limit(self, count: int):
        self.limit_count = count
        return self

    def update(self, payload: dict[str, Any]):
        self.op = "update"
        self.update_payload = payload
        return self

    def insert(self, payload: dict[str, Any] | list[dict[str, Any]]):
        self.op = "insert"
        if isinstance(payload, list):
            self.insert_payload = [deepcopy(item) for item in payload]
        else:
            self.insert_payload = [deepcopy(payload)]
        return self

    def _parse_iso(self, value: Any) -> Any:
        if isinstance(value, datetime):
            return value
        if isinstance(value, str):
            text = value.strip().replace("Z", "+00:00")
            try:
                return datetime.fromisoformat(text)
            except ValueError:
                return value
        return value

    def _matches(self, row: dict[str, Any]) -> bool:
        for op, column, value in self.filters:
            row_value = row.get(column)
            if op == "eq" and row_value != value:
                return False
            if op == "in" and row_value not in value:
                return False
            if op == "is":
                if value == "null" and row_value is not None:
                    return False
                if value != "null" and row_value is not value:
                    return False
            if op == "gte":
                if self._parse_iso(row_value) < self._parse_iso(value):
                    return False
            if op == "lte":
                if self._parse_iso(row_value) > self._parse_iso(value):
                    return False
        return True

    def execute(self) -> FakeResponse:
        table_rows = self._rows()

        if self.op == "insert":
            inserted: list[dict[str, Any]] = []
            for item in self.insert_payload or []:
                if item.get("id") is None:
                    item["id"] = str(uuid.uuid4())
                table_rows.append(item)
                inserted.append(deepcopy(item))
            return FakeResponse(inserted)

        matched = [row for row in table_rows if self._matches(row)]

        if self.op == "update":
            updated: list[dict[str, Any]] = []
            for row in table_rows:
                if not self._matches(row):
                    continue
                for key, value in (self.update_payload or {}).items():
                    row[key] = value
                updated.append(deepcopy(row))
            return FakeResponse(updated)

        if self.order_by is not None:
            key, desc = self.order_by
            matched.sort(key=lambda r: (r.get(key) is None, r.get(key)), reverse=desc)

        if self.range_bounds is not None:
            start, end = self.range_bounds
            matched = matched[start : end + 1]

        if self.limit_count is not None:
            matched = matched[: self.limit_count]

        projected = matched
        if self.select_fields is not None:
            projected = []
            for row in matched:
                projected.append({field: row.get(field) for field in self.select_fields})

        return FakeResponse([deepcopy(item) for item in projected])


class FakeSupabaseClient:
    def __init__(self, db: dict[str, list[dict[str, Any]]]):
        self.db = db

    def table(self, table_name: str) -> FakeQuery:
        return FakeQuery(self.db, table_name)


def assert_true(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def run_smoke_tests() -> None:
    admin_id = "11111111-1111-1111-1111-111111111111"
    student_1_id = "22222222-2222-2222-2222-222222222222"
    student_2_id = "33333333-3333-3333-3333-333333333333"
    intervention_id = "44444444-4444-4444-4444-444444444444"
    notification_id = "55555555-5555-5555-5555-555555555555"

    now = datetime.now(timezone.utc)
    db: dict[str, list[dict[str, Any]]] = {
        "profiles": [
            {
                "id": admin_id,
                "full_name": "TPC Admin",
                "email": "admin@example.com",
                "role": "tpc_admin",
                "department": "CSE",
                "batch_year": 2025,
            },
            {
                "id": student_1_id,
                "full_name": "Riya Sharma",
                "email": "riya@example.com",
                "role": "student",
                "department": "CSE",
                "batch_year": 2025,
            },
            {
                "id": student_2_id,
                "full_name": "Arjun Patel",
                "email": "arjun@example.com",
                "role": "student",
                "department": "ECE",
                "batch_year": 2025,
            },
        ],
        "student_profiles": [
            {
                "id": student_1_id,
                "cgpa": 7.2,
                "active_backlogs": 1,
                "internship_count": 0,
                "mock_tests_attempted": 2,
                "mock_avg_score": 48.5,
                "certifications_count": 1,
                "placement_status": "unplaced",
                "company_placed": None,
                "last_portal_login": (now - timedelta(days=2)).isoformat(),
                "resume_updated_at": (now - timedelta(days=7)).isoformat(),
            },
            {
                "id": student_2_id,
                "cgpa": 8.8,
                "active_backlogs": 0,
                "internship_count": 1,
                "mock_tests_attempted": 5,
                "mock_avg_score": 78.0,
                "certifications_count": 3,
                "placement_status": "placed",
                "company_placed": "Acme Corp",
                "last_portal_login": (now - timedelta(days=1)).isoformat(),
                "resume_updated_at": (now - timedelta(days=3)).isoformat(),
            },
        ],
        "vigilo_scores": [
            {
                "id": str(uuid.uuid4()),
                "student_id": student_1_id,
                "score": 30.0,
                "cluster": "at_risk",
                "placement_probability": 0.32,
                "is_latest": True,
                "computed_at": now.isoformat(),
            },
            {
                "id": str(uuid.uuid4()),
                "student_id": student_2_id,
                "score": 74.0,
                "cluster": "placement_ready",
                "placement_probability": 0.79,
                "is_latest": True,
                "computed_at": now.isoformat(),
            },
        ],
        "interventions": [
            {
                "id": intervention_id,
                "student_id": student_1_id,
                "created_by": admin_id,
                "intervention_type": "weekly_check_in",
                "ai_generated_message": "Please check in with your mentor this week.",
                "custom_message": None,
                "status": "pending",
                "created_at": now.isoformat(),
                "sent_at": None,
                "acknowledged_at": None,
                "deleted_at": None,
            }
        ],
        "alerts": [
            {
                "id": str(uuid.uuid4()),
                "student_id": student_1_id,
                "alert_type": "score_drop",
                "severity": "high",
                "message": "Score dropped significantly",
                "triggered_at": now.isoformat(),
                "is_resolved": True,
                "resolved_at": now.isoformat(),
                "resolved_by": admin_id,
            }
        ],
        "placement_drives": [
            {
                "id": str(uuid.uuid4()),
                "company_name": "Acme Corp",
                "package_lpa": 8.5,
            }
        ],
        "notification_log": [
            {
                "id": notification_id,
                "intervention_id": None,
                "student_id": student_1_id,
                "channel": "in_app",
                "status": "delivered",
                "message_preview": "Initial notification",
                "is_read": False,
                "sent_at": (now - timedelta(hours=1)).isoformat(),
                "delivered_at": (now - timedelta(hours=1)).isoformat(),
                "failed_reason": None,
                "created_at": (now - timedelta(hours=1)).isoformat(),
            }
        ],
    }

    fake_client = FakeSupabaseClient(db)

    notifications_router.get_supabase_client = lambda: fake_client
    interventions_router.get_supabase_client = lambda: fake_client
    interventions_router.broadcast_event = lambda *args, **kwargs: None
    ai_router.get_supabase_client = lambda: fake_client
    ai_router.generate_bulk_nudges = lambda student_ids, intervention_type: [
        {"student_id": sid, "nudge_message": f"{intervention_type} message for {sid}"}
        for sid in student_ids
    ]
    exports_router.get_supabase_client = lambda: fake_client

    app.dependency_overrides[require_tpc_admin] = lambda: {"id": admin_id, "role": "tpc_admin"}
    app.dependency_overrides[get_current_user] = lambda: {"id": admin_id, "role": "tpc_admin"}
    app.dependency_overrides[require_student] = lambda: {"id": student_1_id, "role": "student"}

    with TestClient(app) as client:
        response = client.get(f"/api/v1/notifications/student/{student_1_id}")
        assert_true(response.status_code == 200, "notifications list endpoint failed")

        response = client.get(f"/api/v1/notifications/student/{student_1_id}/unread-count")
        assert_true(response.status_code == 200, "notifications unread-count endpoint failed")
        assert_true(response.json()["data"]["unread"] == 1, "unexpected unread count before read")

        response = client.patch(f"/api/v1/notifications/{notification_id}/read")
        assert_true(response.status_code == 200, "mark notification read failed")

        response = client.get(f"/api/v1/notifications/student/{student_1_id}/unread-count")
        assert_true(response.status_code == 200, "notifications unread-count endpoint failed after read")
        assert_true(response.json()["data"]["unread"] == 0, "unread count should be zero after marking read")

        response = client.get("/api/v1/notifications/feed")
        assert_true(response.status_code == 200, "notifications feed endpoint failed")
        assert_true(len(response.json()["data"]) >= 1, "notifications feed should include at least one row")

        response = client.patch(f"/api/v1/interventions/{intervention_id}/send")
        assert_true(response.status_code == 200, "intervention send endpoint failed")
        sent_notifications = [
            row for row in db["notification_log"] if str(row.get("intervention_id") or "") == intervention_id
        ]
        assert_true(len(sent_notifications) == 1, "intervention send should create one notification row")
        assert_true(sent_notifications[0]["status"] == "delivered", "sent notification should be delivered")

        response = client.post(
            "/api/v1/ai/nudge/bulk",
            json={
                "student_ids": [student_1_id, student_2_id],
                "intervention_type": "weekly_check_in",
            },
        )
        assert_true(response.status_code == 200, "bulk nudge endpoint failed")
        bulk_notifications = [
            row
            for row in db["notification_log"]
            if str(row.get("message_preview") or "").startswith("weekly_check_in")
        ]
        assert_true(len(bulk_notifications) == 2, "bulk nudge should create two notification rows")
        assert_true(all(row.get("status") == "delivered" for row in bulk_notifications), "bulk notifications not delivered")

        response = client.get("/api/v1/exports/students?format=csv")
        assert_true(response.status_code == 200, "students csv export failed")
        assert_true(response.content.startswith(b"\xef\xbb\xbf"), "students csv must include UTF-8 BOM")
        assert_true(response.headers.get("content-type", "").startswith("text/csv"), "students csv content type mismatch")

        response = client.get("/api/v1/exports/students?format=xlsx")
        assert_true(response.status_code == 200, "students xlsx export failed")
        workbook = load_workbook(io.BytesIO(response.content))
        ws = workbook.active
        assert_true(ws.title == "Students", "students xlsx sheet title mismatch")
        assert_true(ws["A1"].value == "Full Name", "students xlsx header mismatch")
        assert_true(ws["K2"].fill.start_color.rgb == "FFFFCCCC", "score conditional formatting mismatch")

        response = client.get("/api/v1/exports/interventions?format=csv")
        assert_true(response.status_code == 200, "interventions csv export failed")

        response = client.get("/api/v1/exports/alerts?format=xlsx")
        assert_true(response.status_code == 200, "alerts xlsx export failed")
        workbook = load_workbook(io.BytesIO(response.content))
        ws = workbook.active
        assert_true(ws.title == "Alerts", "alerts xlsx sheet title mismatch")
        assert_true(ws["D2"].fill.start_color.rgb == "FFFFD8B1", "alert severity fill mismatch")

        response = client.get("/api/v1/exports/placement-summary")
        assert_true(response.status_code == 200, "placement summary export failed")
        workbook = load_workbook(io.BytesIO(response.content))
        assert_true(
            workbook.sheetnames == ["Overview", "By Department", "By Company", "At Risk Students"],
            "placement summary sheet names mismatch",
        )

    print("ALL_CHECKS_PASSED")


if __name__ == "__main__":
    run_smoke_tests()
